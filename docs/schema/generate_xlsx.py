"""
Generates docs/schema/glassbox_schema.xlsx — a color-coded, multi-sheet workbook
containing the full Glass Box AI / AI Claim Processing schema and ER relationships.

Run:  python docs/schema/generate_xlsx.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# ============ COLOR PALETTE ============
COLORS = {
    "header_blue":  "0078D4",   # Azure brand
    "header_purple":"742774",   # Power Platform
    "header_audit": "D99416",   # Audit / compliance
    "pk":           "FFE082",   # Primary key (yellow)
    "fk":           "FFCC80",   # Foreign key (orange)
    "json":         "B3E5FC",   # JSON-typed column (cyan)
    "audit_row":    "FFF3CD",   # audit-critical column (light gold)
    "required":     "C8E6C9",   # Required (light green)
    "computed":     "F5F5F5",   # Computed / system-generated (gray)
    "row_alt":      "F8FAFC",   # Alternating row striping
    "white":        "FFFFFF",
}

WHITE = Font(color="FFFFFF", bold=True, size=12)
BOLD  = Font(bold=True)
ITAL  = Font(italic=True, color="555555")
THIN  = Side(border_style="thin", color="CCCCCC")
BOX   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ============ SCHEMA DATA ============
# (table_name, header_color_key, [columns])
# each column: (logical, physical, type, required, choice/default, lookup, description, tag)
# tag: 'pk' | 'fk' | 'json' | 'audit' | 'required' | None

SCHEMA = [
    ("Policy", "header_blue", [
        ("id",                  "gbx_policy_id",      "Auto-number",          "Y", "POL-{YYYY}-{####}", "",                "Primary key",                                          "pk"),
        ("holder_name",         "gbx_holder_name",    "Text(255)",            "Y", "",                  "",                "Policyholder full name",                               "required"),
        ("holder_email",        "gbx_holder_email",   "Email",                "Y", "",                  "",                "",                                                     "required"),
        ("holder_phone",        "gbx_holder_phone",   "Phone",                "Y", "",                  "",                "",                                                     "required"),
        ("policy_type",         "gbx_policy_type",    "Choice",               "Y", "Auto",              "",                "Personal Auto product line only",                      "required"),
        ("state",               "gbx_state",          "Choice (50 states)",   "Y", "",                  "",                "Drives no-fault routing on related claims",            "required"),
        ("coverage_details",    "gbx_coverage_details","Multiline",           "Y", "",                  "",                "US shorthand 100/300/50 + per-coverage breakdown",     "required"),
        ("total_limit",         "gbx_total_limit",    "Currency",             "N", "",                  "",                "Aggregate liability cap",                              None),
        ("deductible",          "gbx_deductible",     "Currency",             "Y", "",                  "",                "",                                                     "required"),
        ("start_date",          "gbx_start_date",     "Date",                 "Y", "",                  "",                "",                                                     "required"),
        ("end_date",            "gbx_end_date",       "Date",                 "Y", "",                  "",                "",                                                     "required"),
        ("status",              "gbx_status",         "Choice",               "Y", "Active|Lapsed|Cancelled|Pending", "",   "",                                                     "required"),
        ("policy_pdf_url",      "gbx_policy_pdf_url", "URL",                  "N", "",                  "",                "Azure Blob path indexed by AI Search",                 None),
        ("vin",                 "gbx_vin",            "Text(17)",             "N", "",                  "",                "Primary insured vehicle VIN",                          None),
    ]),
    ("Claim", "header_blue", [
        ("id",                      "gbx_claim_id",          "Auto-number",          "Y", "CLM-{YYYY}-{####}", "",                "Primary key",                                       "pk"),
        ("policy_id",               "gbx_policy",            "Lookup",               "Y", "",                  "Policy.id",        "Owning policy",                                     "fk"),
        ("parent_claim_id",         "gbx_parent_claim",      "Lookup",               "N", "",                  "Claim.id",         "Self-reference for Reopen process",                 "fk"),
        ("assigned_adjuster_id",    "gbx_assigned_adjuster", "Lookup",               "N", "",                  "Adjuster.id",      "Set by Handler Assignment Engine",                  "fk"),
        ("channel",                 "gbx_channel",           "Choice",               "Y", "MobileApp|Web|Teams|Email|SMS|WhatsApp", "",  "",                                              "required"),
        ("loss_type",               "gbx_loss_type",         "Choice",               "Y", "Collision|Comp-Weather|Comp-Theft|Comp-Vandalism|Comp-Fire|Comp-Animal|Comp-Glass|Liab-PD|Liab-BI|PIP-MedPay|UM-UIM", "",   "11 Personal Auto loss types",                       "required"),
        ("sub_type",                "gbx_sub_type",          "Text(100)",            "N", "",                  "",                 "e.g. Parked-and-struck, Hail, Hit-and-run",         None),
        ("incident_date",           "gbx_incident_date",     "DateTime",             "Y", "",                  "",                 "",                                                  "required"),
        ("incident_state",          "gbx_incident_state",    "Choice (50 states)",   "Y", "",                  "",                 "Drives no-fault routing",                           "required"),
        ("location",                "gbx_location",          "Text(500)",            "Y", "",                  "",                 "Address or {lat;lng}",                              "required"),
        ("description",             "gbx_description",       "Multiline",            "Y", "",                  "",                 "U6 narrative — fed to LLM + sentiment check",       "required"),
        ("loss_type_details",       "gbx_loss_type_details", "Multiline (JSON)",     "Y", "",                  "",                 "All loss-type-specific answers as JSON. Shape varies per loss_type.", "json"),
        ("vin",                     "gbx_vin",               "Text(17)",             "N", "",                  "",                 "",                                                  None),
        ("police_report_number",    "gbx_police_report_number","Text(50)",           "N", "",                  "",                 "If U8=Y",                                           None),
        ("responding_agency",       "gbx_responding_agency", "Text(200)",            "N", "",                  "",                 "If U8=Y",                                           None),
        ("injury_flag",             "gbx_injury_flag",       "Yes/No",               "Y", "",                  "",                 "U7",                                                "required"),
        ("distress_flag",           "gbx_distress_flag",     "Yes/No",               "N", "",                  "",                 "Set by sentiment analysis",                         None),
        ("vehicle_drivable",        "gbx_vehicle_drivable",  "Yes/No",               "N", "",                  "",                 "",                                                  None),
        ("vehicle_towed",           "gbx_vehicle_towed",     "Yes/No",               "N", "",                  "",                 "",                                                  None),
        ("tow_location",            "gbx_tow_location",      "Text(255)",            "N", "",                  "",                 "",                                                  None),
        ("other_party_involved",    "gbx_other_party_involved","Yes/No",             "Y", "",                  "",                 "",                                                  "required"),
        ("status",                  "gbx_status",            "Choice",               "Y", "New|Processing|AwaitingDocs|UnderReview|Approved|Denied|Escalated|Cancelled|Reopen", "",  "",                                                  "required"),
        ("confidence_score",        "gbx_confidence_score",  "Whole number (0-100)", "N", "",                  "",                 "Set by Adjudication Agent",                         None),
        ("recommendation",          "gbx_recommendation",    "Choice",               "N", "Approve|Deny|Partial|Escalate|Adjust", "", "",                                              None),
        ("settlement_amount",       "gbx_settlement_amount", "Currency",             "N", "",                  "",                 "Final payout",                                      None),
        ("settlement_type",         "gbx_settlement_type",   "Choice",               "N", "DRP|Cash|TotalLoss|DirectToThirdParty", "", "",                                            None),
        ("tier",                    "gbx_tier",              "Choice",               "N", "1|2|3",             "",                 "Routing tier from Adjudication",                    None),
        ("assignment_reason",       "gbx_assignment_reason", "Multiline",            "N", "",                  "",                 "Plain-English why this adjuster",                   None),
        ("assigned_at",             "gbx_assigned_at",       "DateTime",             "N", "",                  "",                 "Starts the SLA clock",                              None),
        ("sla_due_at",              "gbx_sla_due_at",        "DateTime",             "N", "",                  "",                 "Computed at assignment time per tier",              None),
        ("resolved_at",             "gbx_resolved_at",       "DateTime",             "N", "",                  "",                 "",                                                  None),
    ]),
    ("Document", "header_blue", [
        ("id",                  "gbx_doc_id",         "Auto-number",          "Y", "DOC-{YYYY}-{#####}", "",            "Primary key",                                          "pk"),
        ("claim_id",            "gbx_claim",          "Lookup",               "Y", "",                   "Claim.id",     "Owning claim",                                         "fk"),
        ("doc_type",            "gbx_doc_type",       "Choice",               "Y", "DamagePhoto|DriversLicense|InsuranceCard|VehicleTitle|RegistrationCard|RepairEstimate|PoliceReport|FireDeptReport|MedicalRecord|LostWagesLetter|ContractorEstimate|WitnessStatement|Other", "", "13 document types",                "required"),
        ("file_name",           "gbx_file_name",      "Text(255)",            "Y", "",                   "",             "",                                                     "required"),
        ("file_url",            "gbx_file_url",       "URL",                  "Y", "",                   "",             "Azure Blob path",                                      "required"),
        ("mime_type",           "gbx_mime_type",      "Text(50)",             "N", "",                   "",             "",                                                     None),
        ("size_bytes",          "gbx_size_bytes",     "Whole number",         "N", "",                   "",             "",                                                     None),
        ("extracted_data",      "gbx_extracted_data", "Multiline (JSON)",     "N", "",                   "",             "Doc Intelligence response, parsed",                    "json"),
        ("quality_score",       "gbx_quality_score",  "Decimal (0-1)",        "N", "",                   "",             "From GPT-4o Vision photo-quality check",               None),
        ("missing_fields",      "gbx_missing_fields", "Multiline (JSON arr)", "N", "",                   "",             "If extraction couldn't find required fields",          "json"),
        ("upload_date",         "gbx_upload_date",    "DateTime",             "Y", "",                   "",             "",                                                     "required"),
        ("channel",             "gbx_channel",        "Choice",               "Y", "(same as Claim.channel)", "",        "",                                                     "required"),
    ]),
    ("Communication", "header_blue", [
        ("id",                  "gbx_comm_id",        "Auto-number",          "Y", "COM-{YYYY}-{######}", "",            "Primary key",                                          "pk"),
        ("claim_id",            "gbx_claim",          "Lookup",               "Y", "",                   "Claim.id",     "Owning claim",                                         "fk"),
        ("direction",           "gbx_direction",      "Choice",               "Y", "Inbound|Outbound",   "",             "",                                                     "required"),
        ("channel",             "gbx_channel",        "Choice",               "Y", "(same as Claim.channel)", "",        "",                                                     "required"),
        ("message_content",     "gbx_message_content","Multiline",            "Y", "",                   "",             "",                                                     "required"),
        ("attachments",         "gbx_attachments",    "Multiline (JSON arr)", "N", "",                   "",             "Array of {name; url; size}",                           "json"),
        ("timestamp",           "gbx_timestamp",      "DateTime",             "Y", "",                   "",             "",                                                     "required"),
        ("agent_name",          "gbx_agent_name",     "Text(100)",            "N", "",                   "",             "Intake | Explanation | Adjuster | Notification",       None),
        ("sentiment_score",     "gbx_sentiment_score","Decimal (-1 to 1)",    "N", "",                   "",             "From AOAI on inbound only",                            None),
        ("sender_id",           "gbx_sender_id",      "Text(255)",            "N", "",                   "",             "",                                                     None),
    ]),
    ("Decision_Rationale", "header_audit", [
        ("id",                          "gbx_log_id",                       "Auto-number",          "Y", "LOG-{YYYY}-{#######}", "",       "Primary key",                                          "pk"),
        ("claim_id",                    "gbx_claim",                        "Lookup",               "Y", "",                     "Claim.id","Owning claim",                                         "fk"),
        ("agent_name",                  "gbx_agent_name",                   "Choice",               "Y", "Intake|Extraction|Policy|Validation|Adjudication|Explanation|Notification|Adjuster|AssignmentEngine|VendorEngine", "", "Which agent wrote this row",                          "required"),
        ("sub_agent",                   "gbx_sub_agent",                    "Text(50)",             "N", "",                     "",       "NOAA|NHTSA|ISO|NICB|DMV|Telematics|EstimateRule|Sentiment", None),
        ("action",                      "gbx_action",                       "Text(255)",            "Y", "",                     "",       "Short description of the step",                        "required"),
        ("policy_reference",            "gbx_policy_reference",             "Text(255)",            "N", "",                     "",       "e.g. 'Section 4.2 — Collision Coverage'",              None),
        ("data_points",                 "gbx_data_points",                  "Multiline (JSON)",     "N", "",                     "",       "Input the agent saw",                                  "json"),
        ("confidence_contribution",     "gbx_confidence_contribution",      "Decimal",              "N", "",                     "",       "How this step affected overall confidence (-1 to +1)", None),
        ("external_api_result",         "gbx_external_api_result",          "Multiline (JSON)",     "N", "",                     "",       "Raw response from any external call",                  "json"),
        ("adapter_status",              "gbx_adapter_status",               "Choice",               "N", "Live|Sandbox|NotApplicable", "", "",                                                     None),
        ("flag_raised",                 "gbx_flag_raised",                  "Yes/No",               "N", "",                     "",       "True if step raised a fraud or escalation flag",       None),
        ("flag_severity",               "gbx_flag_severity",                "Choice",               "N", "Low|Medium|High",      "",       "",                                                     None),
        ("human_readable_explanation",  "gbx_human_readable_explanation",   "Multiline",            "Y", "",                     "",       "⭐ THE compliance text. Plain-English sentence regulators read.", "audit"),
        ("latency_ms",                  "gbx_latency_ms",                   "Whole number",         "N", "",                     "",       "",                                                     None),
        ("timestamp",                   "gbx_timestamp",                    "DateTime",             "Y", "",                     "",       "",                                                     "required"),
    ]),
    ("Adjuster", "header_purple", [
        ("id",                      "gbx_adjuster_id",     "Auto-number",          "Y", "ADJ-{####}",       "",       "Primary key",                                          "pk"),
        ("aad_email",               "gbx_aad_email",       "Email (unique)",       "Y", "",                 "",       "Microsoft Entra ID email — SSO identity",              "required"),
        ("display_name",            "gbx_display_name",    "Text(200)",            "Y", "",                 "",       "",                                                     "required"),
        ("seniority",               "gbx_seniority",       "Choice",               "Y", "Junior|Senior|Lead", "",     "",                                                     "required"),
        ("skills",                  "gbx_skills",          "Choice (multi)",       "Y", "Auto-Collision|Auto-Comprehensive|Auto-Liability|BI|PIP|UM-Specialist|SIU-Fraud|Property-knowledge", "", "Used for skills-based routing", "required"),
        ("state_licenses",          "gbx_state_licenses",  "Choice (multi 50)",    "Y", "",                 "",       "Adjuster must be licensed in claim's incident_state", "required"),
        ("is_available",            "gbx_is_available",    "Yes/No",               "Y", "",                 "",       "Currently on-shift and accepting new assignments",     "required"),
        ("max_workload_per_day",    "gbx_max_workload",    "Whole number",         "Y", "15 default",       "",       "Cap on concurrent active claims",                      "required"),
        ("current_workload",        "gbx_current_workload","Whole number",         "N", "computed",         "",       "Computed from active Claim assignments",               "computed"),
        ("timezone",                "gbx_timezone",        "Choice",               "N", "US/Pacific|US/Mountain|US/Central|US/Eastern", "", "Timezone-alignment scoring", None),
        ("excluded_policy_ids",     "gbx_excluded_policies","Multiline",           "N", "",                 "",       "Conflict-of-interest exclusions",                      None),
    ]),
    ("Vendor", "header_purple", [
        ("id",                      "gbx_vendor_id",        "Auto-number",         "Y", "VND-{#####}",      "",       "Primary key",                                          "pk"),
        ("vendor_type",             "gbx_vendor_type",      "Choice",              "Y", "DRPShop|TowingCompany|GlassShop|RentalPartner|SalvageYard|IndependentAppraiser|MedicalProvider", "", "7 vendor types", "required"),
        ("name",                    "gbx_name",             "Text(200)",           "Y", "",                 "",       "Business name",                                        "required"),
        ("address",                 "gbx_address",          "Text(500)",           "Y", "",                 "",       "Street + city + state + ZIP",                          "required"),
        ("latitude",                "gbx_lat",              "Decimal",             "Y", "",                 "",       "For distance-from-incident scoring",                   "required"),
        ("longitude",               "gbx_lng",              "Decimal",             "Y", "",                 "",       "",                                                     "required"),
        ("service_radius_miles",    "gbx_service_radius",   "Whole number",        "Y", "25 default",       "",       "Max distance from base they will travel",              "required"),
        ("services_offered",        "gbx_services",         "Choice (multi)",      "Y", "(varies by vendor_type)", "", "e.g. Auto-Body | Glass-Replace | Tow-Standard",        "required"),
        ("is_active_in_drp",        "gbx_in_drp",           "Yes/No",              "Y", "",                 "",       "Part of carrier's preferred network",                  "required"),
        ("customer_rating",         "gbx_customer_rating",  "Decimal (0-5)",       "N", "",                 "",       "Average customer rating",                              None),
        ("drp_performance_score",   "gbx_drp_score",        "Decimal (0-100)",     "N", "",                 "",       "Carrier-internal: turnaround time + CSAT composite",   None),
        ("current_capacity",        "gbx_capacity",         "Whole number",        "N", "computed",         "",       "Available capacity slots",                             "computed"),
        ("phone",                   "gbx_vendor_phone",     "Phone",               "Y", "",                 "",       "",                                                     "required"),
        ("email",                   "gbx_vendor_email",     "Email",               "N", "",                 "",       "",                                                     None),
        ("states_served",           "gbx_states",           "Choice (multi 50)",   "Y", "",                 "",       "",                                                     "required"),
    ]),
    ("ClaimVendorAssignment", "header_audit", [
        ("id",                      "gbx_cva_id",          "Auto-number",          "Y", "CVA-{######}",     "",        "Primary key",                                          "pk"),
        ("claim_id",                "gbx_claim",           "Lookup",               "Y", "",                 "Claim.id", "",                                                     "fk"),
        ("vendor_id",               "gbx_vendor",          "Lookup",               "Y", "",                 "Vendor.id","",                                                     "fk"),
        ("assignment_purpose",      "gbx_purpose",         "Choice",               "Y", "Repair|Tow|Glass|Rental|Salvage|Inspect|Medical", "", "Why this vendor for this claim", "required"),
        ("distance_miles",          "gbx_distance",        "Decimal",              "N", "",                 "",         "Computed at assignment time",                          None),
        ("score",                   "gbx_score",           "Decimal (0-100)",      "N", "",                 "",         "Routing score — logged for audit",                     "audit"),
        ("assignment_reason",       "gbx_reason",          "Multiline",            "N", "",                 "",         "Plain-English why this vendor",                        "audit"),
        ("assigned_at",             "gbx_assigned_at",     "DateTime",             "Y", "",                 "",         "",                                                     "required"),
        ("sla_due_at",              "gbx_sla_due_at",      "DateTime",             "N", "",                 "",         "",                                                     None),
        ("status",                  "gbx_cva_status",      "Choice",               "Y", "Notified|Accepted|InProgress|Completed|Cancelled|Declined", "", "",                "required"),
        ("completion_notes",        "gbx_completion_notes","Multiline",            "N", "",                 "",         "Vendor's update on completion",                        None),
    ]),
]

RELATIONSHIPS = [
    # (from_table, from_col, to_table, to_col, cardinality, cascade, notes)
    ("Claim",                    "policy_id",           "Policy",  "id", "N:1", "Restrict",          "Every claim belongs to exactly one policy"),
    ("Claim",                    "parent_claim_id",     "Claim",   "id", "N:1", "Set Null",          "Self-reference for Reopen process"),
    ("Claim",                    "assigned_adjuster_id","Adjuster","id", "N:1", "Set Null",          "Set by Handler Assignment Engine"),
    ("Document",                 "claim_id",            "Claim",   "id", "N:1", "Cascade Delete",    "Documents are owned by the claim"),
    ("Communication",            "claim_id",            "Claim",   "id", "N:1", "Cascade Delete",    "Communications belong to the claim"),
    ("Decision_Rationale",       "claim_id",            "Claim",   "id", "N:1", "Cascade Delete",    "Audit log entries belong to the claim"),
    ("ClaimVendorAssignment",    "claim_id",            "Claim",   "id", "N:1", "Cascade Delete",    "CVA links a claim to vendors"),
    ("ClaimVendorAssignment",    "vendor_id",           "Vendor",  "id", "N:1", "Restrict",          "Don't delete vendors with active assignments"),
]

def fill(c):
    return PatternFill(start_color=c, end_color=c, fill_type="solid")

def style_header(cell, color_key="header_blue"):
    cell.fill = fill(COLORS[color_key])
    cell.font = WHITE
    cell.alignment = CENTER
    cell.border = BOX

def style_subheader(cell):
    cell.fill = fill("E2E8F0")
    cell.font = BOLD
    cell.alignment = CENTER
    cell.border = BOX

def style_data(cell, tag=None, alt=False):
    if tag == "pk":
        cell.fill = fill(COLORS["pk"])
        cell.font = BOLD
    elif tag == "fk":
        cell.fill = fill(COLORS["fk"])
        cell.font = BOLD
    elif tag == "json":
        cell.fill = fill(COLORS["json"])
    elif tag == "audit":
        cell.fill = fill(COLORS["audit_row"])
        cell.font = BOLD
    elif tag == "required":
        cell.fill = fill(COLORS["required"])
    elif tag == "computed":
        cell.fill = fill(COLORS["computed"])
        cell.font = ITAL
    elif alt:
        cell.fill = fill(COLORS["row_alt"])
    cell.alignment = LEFT
    cell.border = BOX

def build_overview_sheet(wb):
    ws = wb.create_sheet("Overview", 0)
    ws.cell(1, 1, "AI Claim Processing — Schema Overview").font = Font(bold=True, size=16, color="0a2540")
    ws.cell(2, 1, "8 entities · Microsoft Dataverse · color-coded for PK / FK / JSON / audit / required").font = ITAL

    headers = ["Table", "Category", "Columns", "Has FK?", "Has JSON?", "Audit-critical?", "Purpose"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(4, col, h)
        style_header(c, "header_blue")

    rows = [
        ("Policy",                  "Core",              "14", "—",  "—",  "—",  "Customer's insurance policy. Loaded from sample data; never created at runtime."),
        ("Claim",                   "Core (hub)",        "31", "✓",  "✓",  "—",  "The central entity. Created at FNOL submission. Updated by every agent."),
        ("Document",                "Core",              "12", "✓",  "✓",  "—",  "Uploaded files (photos, IDs, estimates, etc.). Extracted data stored as JSON."),
        ("Communication",           "Core",              "10", "✓",  "✓",  "—",  "Inbound + outbound messages per channel."),
        ("Decision_Rationale",      "Audit / compliance","15", "✓",  "✓",  "✓",  "⭐ Glass Box audit log. Every AI step writes one row. THE compliance artifact."),
        ("Adjuster",                "Pool (handlers)",   "11", "—",  "—",  "—",  "Carrier employee pool. Used by Handler Assignment Engine."),
        ("Vendor",                  "Pool (3rd parties)","15", "—",  "—",  "—",  "DRP shops, towing, glass, rental, salvage, etc. Used by Vendor Assignment Engine."),
        ("ClaimVendorAssignment",   "Join + audit",      "11", "✓",  "—",  "✓",  "Many-to-many between Claim and Vendor. One row per vendor assignment with score + reason."),
    ]
    for r_idx, row in enumerate(rows, 5):
        for c_idx, v in enumerate(row, 1):
            c = ws.cell(r_idx, c_idx, v)
            tag = "audit" if r_idx in (9, 12) else "fk" if "✓" in str(row[3]) else None
            style_data(c, tag, alt=(r_idx % 2 == 0))

    # Legend
    ws.cell(15, 1, "Color legend").font = Font(bold=True, size=12)
    legend = [
        ("🔑 Primary key",       "pk"),
        ("🔗 Foreign key",       "fk"),
        ("📦 JSON column",       "json"),
        ("⭐ Audit-critical",     "audit"),
        ("✅ Required",          "required"),
        ("⚙ Computed/system",   "computed"),
    ]
    for i, (label, tag) in enumerate(legend):
        c = ws.cell(16 + i, 1, label)
        style_data(c, tag)

    # Column widths
    widths = [28, 22, 10, 10, 12, 16, 70]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 24
    ws.freeze_panes = "A5"

def build_table_sheet(wb, name, header_color, columns):
    ws = wb.create_sheet(name)
    ws.cell(1, 1, name).font = Font(bold=True, size=18, color=COLORS[header_color])
    ws.cell(2, 1, f"Logical column names shown · physical names use gbx_ prefix (Dataverse)").font = ITAL

    headers = ["Column (logical)", "Physical (Dataverse)", "Type", "Required", "Choice / Default", "Lookup →", "Description"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(4, col, h)
        style_header(c, header_color)

    for r_idx, col in enumerate(columns, 5):
        logical, physical, dtype, required, choice, lookup, desc, tag = col
        values = [logical, physical, dtype, required, choice, lookup, desc]
        for c_idx, v in enumerate(values, 1):
            c = ws.cell(r_idx, c_idx, v)
            style_data(c, tag, alt=(r_idx % 2 == 0))

    widths = [28, 32, 22, 10, 50, 18, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 28
    ws.freeze_panes = "A5"

def build_relationships_sheet(wb):
    ws = wb.create_sheet("Relationships")
    ws.cell(1, 1, "Foreign-key relationships").font = Font(bold=True, size=16, color="0a2540")
    ws.cell(2, 1, "8 relationships · all Dataverse Lookups · cascade behavior tells you what happens when the parent is deleted").font = ITAL

    headers = ["From Table", "From Column", "→", "To Table", "To Column", "Cardinality", "Cascade", "Notes"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(4, col, h)
        style_header(c, "header_blue")

    for r_idx, rel in enumerate(RELATIONSHIPS, 5):
        from_t, from_c, to_t, to_c, card, casc, notes = rel
        values = [from_t, from_c, "→", to_t, to_c, card, casc, notes]
        for c_idx, v in enumerate(values, 1):
            c = ws.cell(r_idx, c_idx, v)
            tag = "audit" if "Cascade" in casc else None
            style_data(c, tag, alt=(r_idx % 2 == 0))
            if c_idx == 3:
                c.alignment = CENTER
                c.font = Font(bold=True, size=14, color="0078D4")

    widths = [24, 24, 4, 24, 14, 14, 18, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 24
    ws.freeze_panes = "A5"

def main():
    wb = Workbook()
    wb.remove(wb.active)
    build_overview_sheet(wb)
    for name, header_color, cols in SCHEMA:
        build_table_sheet(wb, name, header_color, cols)
    build_relationships_sheet(wb)

    out = Path(__file__).parent / "glassbox_schema.xlsx"
    wb.save(out)
    print(f"Wrote {out} with {len(wb.sheetnames)} sheets: {wb.sheetnames}")

if __name__ == "__main__":
    main()
